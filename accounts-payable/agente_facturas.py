"""
AGENTE DE REGISTRO DE FACTURAS DE COMPRA - ISBEROAL
=====================================================
Descripción: Lee correos de contabilidad@isberoal.com, descarga facturas adjuntas,
             extrae datos con OCR + IA, genera CSV para Holded y lo importa via API.

Requisitos:
    pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client
    pip install anthropic python-dotenv pymupdf pillow requests pandas

Configuración:
    Crear archivo .env con:
        HOLDED_API_KEY=tu_api_key_aqui
        GOOGLE_CREDENTIALS_PATH=credentials.json

    Variables de entorno opcionales (modo Railway / producción headless):
        GMAIL_TOKEN_JSON       JSON completo del token Gmail (sustituye a token.json)
        RAILWAY_ENVIRONMENT    Cualquier valor activa modo Railway (logs solo a stdout,
                               sin fallback OAuth interactivo)
        ISBEROAL_SHADOW_MODE   "true" fuerza --solo-xlsx y desactiva import a Holded
                               (modo seguro para validación en paralelo al agente local)
"""

import os
import base64
import json
import re
import csv
import time
import logging
from datetime import datetime, timedelta
from pathlib import Path

import anthropic
import fitz  # PyMuPDF
import pandas as pd
import requests
from dotenv import load_dotenv
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────

load_dotenv(Path(__file__).parent / ".env", override=True)

HOLDED_API_KEY    = os.getenv("HOLDED_API_KEY")
GMAIL_CREDENTIALS = os.getenv("GOOGLE_CREDENTIALS_PATH", "credentials.json")
GMAIL_TOKEN_PATH  = "token.json"

# Scopes de Gmail (solo lectura)
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

# Directorios de trabajo
BASE_DIR         = Path(__file__).parent
FACTURAS_DIR     = BASE_DIR / "Facturas_de_gasto"
CSV_OUTPUT_DIR   = BASE_DIR / "csv_output"
XLSX_OUTPUT_DIR  = BASE_DIR / "xlsx_output"
LOGS_DIR         = BASE_DIR / "logs"
PROCESADOS_PATH  = BASE_DIR / "mensajes_procesados.json"

for d in [FACTURAS_DIR, CSV_OUTPUT_DIR, XLSX_OUTPUT_DIR, LOGS_DIR]:
    d.mkdir(exist_ok=True)

# Logging
# En Railway (sistema de archivos efímero) solo escribimos a stdout.
# En local mantenemos el FileHandler para tener histórico en disco.
_log_handlers = [logging.StreamHandler()]
if not os.getenv("RAILWAY_ENVIRONMENT"):
    _log_handlers.insert(0, logging.FileHandler(LOGS_DIR / f"agente_{datetime.now().strftime('%Y%m%d')}.log"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=_log_handlers
)
log = logging.getLogger(__name__)

# Cliente Anthropic para OCR y extracción
claude = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

# ─────────────────────────────────────────────
# PASO 1 - CONEXIÓN A GMAIL
# ─────────────────────────────────────────────

def conectar_gmail():
    """
    Autentica con Gmail y devuelve el servicio.

    Orden de prioridad:
    1. Variable de entorno GMAIL_TOKEN_JSON (modo Railway/producción headless)
    2. Archivo token.json local (modo desarrollo en máquina con navegador)
    3. Flujo OAuth interactivo con navegador (primera vez en local)
    """
    creds = None

    token_env = os.getenv("GMAIL_TOKEN_JSON")
    if token_env:
        try:
            creds = Credentials.from_authorized_user_info(json.loads(token_env), SCOPES)
            log.info("[OK] Credenciales Gmail cargadas desde variable de entorno")
        except Exception as e:
            log.error(f"[ERROR] No se pudo parsear GMAIL_TOKEN_JSON: {e}")
            creds = None

    if not creds and Path(GMAIL_TOKEN_PATH).exists():
        creds = Credentials.from_authorized_user_file(GMAIL_TOKEN_PATH, SCOPES)

    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        if not token_env:
            with open(GMAIL_TOKEN_PATH, "w") as f:
                f.write(creds.to_json())

    if not creds or not creds.valid:
        if os.getenv("RAILWAY_ENVIRONMENT"):
            raise RuntimeError(
                "No hay credenciales válidas en GMAIL_TOKEN_JSON. "
                "Genera el token en local y súbelo a Railway como variable de entorno."
            )
        flow = InstalledAppFlow.from_client_secrets_file(GMAIL_CREDENTIALS, SCOPES)
        creds = flow.run_local_server(port=0)
        with open(GMAIL_TOKEN_PATH, "w") as f:
            f.write(creds.to_json())

    service = build("gmail", "v1", credentials=creds)
    log.info("[OK] Conectado a Gmail correctamente")
    return service


def obtener_correos_con_facturas(service, dias_atras=7):
    """
    Busca correos enviados a contabilidad@isberoal.com con adjuntos
    en los últimos N días.
    """
    fecha_desde = (datetime.now() - timedelta(days=dias_atras)).strftime("%Y/%m/%d")
    query = f"to:contabilidad@isberoal.com has:attachment after:{fecha_desde}"

    result = service.users().messages().list(
        userId="me",
        q=query,
        maxResults=50
    ).execute()

    mensajes = result.get("messages", [])
    log.info(f"[CORREO] Encontrados {len(mensajes)} correos con facturas")
    return mensajes


# ─────────────────────────────────────────────
# PASO 2 - DESCARGAR Y RENOMBRAR ADJUNTOS
# ─────────────────────────────────────────────

def descargar_adjuntos(service, mensaje_id):
    """
    Descarga los adjuntos PDF/imagen de un mensaje de Gmail.
    Devuelve lista de rutas descargadas.
    """
    msg = service.users().messages().get(userId="me", id=mensaje_id).execute()
    partes = msg.get("payload", {}).get("parts", [])
    archivos_descargados = []

    # Extraer fecha del correo para el nombre
    headers = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}
    fecha_str = headers.get("Date", "")
    try:
        from email.utils import parsedate_to_datetime
        fecha_email = parsedate_to_datetime(fecha_str)
        fecha_prefix = fecha_email.strftime("%Y%m%d")
    except Exception:
        fecha_prefix = datetime.now().strftime("%Y%m%d")

    for parte in partes:
        filename = parte.get("filename", "")
        mime     = parte.get("mimeType", "")

        if not filename:
            continue

        # Solo procesar PDFs e imágenes
        extensiones_validas = [".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".bmp", ".webp"]
        ext = Path(filename).suffix.lower()
        if ext not in extensiones_validas:
            continue

        body = parte.get("body", {})
        att_id = body.get("attachmentId")
        if not att_id:
            continue

        att = service.users().messages().attachments().get(
            userId="me",
            messageId=mensaje_id,
            id=att_id
        ).execute()

        data = base64.urlsafe_b64decode(att["data"])
        ruta_temp = FACTURAS_DIR / f"temp_{mensaje_id}_{filename}"

        with open(ruta_temp, "wb") as f:
            f.write(data)

        log.info(f"[DESCARGA] {filename}")
        archivos_descargados.append({
            "ruta": ruta_temp,
            "fecha_prefix": fecha_prefix,
            "mime": mime
        })

    return archivos_descargados


# ─────────────────────────────────────────────
# PASO 3 - OCR Y EXTRACCIÓN DE DATOS CON IA
# ─────────────────────────────────────────────

def pdf_a_imagenes(ruta_pdf):
    """Convierte páginas de un PDF a imágenes base64."""
    imagenes = []
    doc = fitz.open(str(ruta_pdf))
    for page in doc:
        pix = page.get_pixmap(dpi=200)
        img_bytes = pix.tobytes("png")
        img_b64 = base64.standard_b64encode(img_bytes).decode("utf-8")
        imagenes.append(img_b64)
    doc.close()
    return imagenes


def imagen_a_base64(ruta_imagen):
    """Convierte una imagen a base64."""
    with open(ruta_imagen, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8")


def extraer_datos_factura(ruta_archivo):
    """
    Usa Claude Vision para extraer datos estructurados de la factura.
    Devuelve un diccionario con los campos de Holded.
    """
    ext = Path(ruta_archivo).suffix.lower()

    # Preparar imágenes para Claude
    if ext == ".pdf":
        imagenes_b64 = pdf_a_imagenes(ruta_archivo)
        media_type = "image/png"
    else:
        imagenes_b64 = [imagen_a_base64(ruta_archivo)]
        media_type = f"image/{ext.strip('.')}"
        if media_type == "image/jpg":
            media_type = "image/jpeg"

    # Construir contenido para Claude
    content = []
    for img_b64 in imagenes_b64[:3]:  # máximo 3 páginas
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": media_type,
                "data": img_b64
            }
        })

    content.append({
        "type": "text",
        "text": """Analiza este documento y extrae los datos en formato JSON estricto.
Responde UNICAMENTE con el JSON, sin texto adicional, sin bloques de codigo markdown.

CONTEXTO CRITICO:
- El COMPRADOR/CLIENTE siempre es ISBEROAL, S.L. con NIF B72598022 (o Isberoal Energy).
- Tu tarea es extraer los datos del VENDEDOR/PROVEEDOR/EMISOR, que es QUIEN EMITE la factura.
- NUNCA uses los datos de ISBEROAL como proveedor. ISBEROAL es siempre el destinatario/comprador.
- Si el NIF que extraes es B72598022, estas extrayendo el comprador por error. Busca el otro NIF.

CLASIFICACION DEL DOCUMENTO:
- Primero determina si es una FACTURA o un ALBARAN/NOTA DE ENTREGA/PRESUPUESTO/RECIBO.
- Un albaran NO tiene base imponible, IVA ni total a pagar. Solo lista productos entregados.
- Un recibo de pago (receipt) NO es una factura, es un comprobante de que ya se pago.

{
  "tipo_documento": "factura | albaran | recibo | presupuesto | otro",
  "es_factura": true,
  "num_factura": "numero de factura",
  "fecha": "DD/MM/YYYY",
  "fecha_vencimiento": "DD/MM/YYYY",
  "proveedor": "nombre completo del VENDEDOR/EMISOR (NO de Isberoal)",
  "nif": "CIF o NIF del VENDEDOR/EMISOR (NO el de Isberoal B72598022)",
  "direccion": "direccion del proveedor/emisor",
  "poblacion": "ciudad del proveedor/emisor",
  "codigo_postal": "codigo postal del proveedor/emisor",
  "provincia": "provincia del proveedor/emisor",
  "pais": "nombre del pais del proveedor/emisor",
  "codigo_pais": "codigo ISO 3166-1 alpha-2 del pais del proveedor (ES, FR, DE, IE, PT, US, CN...)",
  "lineas": [
    {
      "concepto": "descripcion del producto o servicio",
      "precio_unidad": 0.00,
      "unidades": 1,
      "descuento": 0,
      "iva": 21,
      "retencion": 0
    }
  ],
  "base_imponible": 0.00,
  "total_iva": 0.00,
  "total_retencion": 0.00,
  "total_factura": 0.00,
  "moneda": "EUR",
  "inversion_sujeto_pasivo": false
}

Reglas importantes:
- tipo_documento: clasificar el documento. Si NO es una factura, poner es_factura: false
- proveedor y nif: SIEMPRE del VENDEDOR/EMISOR. Busca en la cabecera del documento quien lo emite.
  Normalmente aparece arriba a la izquierda o con etiquetas como "Emisor", "Vendedor", "De:", etc.
  Los datos del comprador suelen estar bajo "Cliente", "Destinatario", "Para:", "Facturar a:", etc.
- fechas siempre en formato DD/MM/YYYY
- precio_unidad sin IVA
- iva como numero entero (21, 10, 4, 0)
- retencion: porcentaje de retencion de IRPF de cada linea (0 si no aplica, 15 si es 15%, etc.)
- total_retencion: suma total de retenciones en euros
- si hay varias lineas de concepto, incluir todas en el array lineas
- si no encuentras fecha de vencimiento, usa 30 dias desde la fecha de emision
- total_factura es el importe total final (base + IVA - retenciones). NUNCA puede ser 0 en una factura real.
- inversion_sujeto_pasivo: true si la factura indica inversion del sujeto pasivo, false en caso contrario
- codigo_pais: SIEMPRE incluir el codigo ISO del pais del proveedor/emisor
  - NIF/CIF que empieza por letra(s) de pais europeo (FR, DE, PT, IT, IE...): usar ese codigo
  - NIF/CIF espanol (empieza por A-H, J, N, P, Q, R, S, U, V, W o un numero): codigo ES
  - Si no puedes determinarlo, usar el pais indicado en la direccion del emisor
"""
    })

    response = claude.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2000,
        messages=[{"role": "user", "content": content}]
    )

    texto = response.content[0].text.strip()

    # Limpiar posibles bloques markdown
    texto = re.sub(r"```json\s*", "", texto)
    texto = re.sub(r"```\s*", "", texto)

    datos = json.loads(texto)
    log.info(f"[OCR] {datos.get('tipo_documento', 'factura').upper()} {datos.get('num_factura')} - {datos.get('proveedor')}")
    return datos


def validar_datos_factura(datos, ruta_archivo):
    """
    Valida los datos extraidos de una factura. Devuelve (valido, motivo).
    """
    # Descartar documentos que no son facturas
    if not datos.get("es_factura", True):
        tipo = datos.get("tipo_documento", "desconocido")
        return False, f"No es factura, es {tipo}"

    # Verificar que el NIF no sea el de ISBEROAL (confundio comprador con vendedor)
    nif = datos.get("nif", "").replace("-", "").replace(" ", "").strip()
    if nif == NIF_ISBEROAL:
        return False, f"NIF extraido es el de ISBEROAL ({NIF_ISBEROAL}), confundio comprador con vendedor"

    # Verificar que hay total > 0
    total = datos.get("total_factura", 0)
    if not total or float(total) <= 0:
        return False, f"Total factura es {total} (cero o negativo)"

    # Verificar que hay proveedor
    proveedor = datos.get("proveedor", "").strip()
    if not proveedor:
        return False, "Sin nombre de proveedor"

    return True, "OK"


# ─────────────────────────────────────────────
# PASO 2 (continuación) - RENOMBRAR ARCHIVO
# ─────────────────────────────────────────────

def renombrar_factura(ruta_temp, datos):
    """
    Renombra el archivo con el formato: AAAAMMDD_FG_Proveedor_descripcion.ext
    """
    ext = ruta_temp.suffix.lower()

    # Fecha
    try:
        fecha_obj = datetime.strptime(datos["fecha"], "%d/%m/%Y")
        fecha_str = fecha_obj.strftime("%Y%m%d")
    except Exception:
        fecha_str = datetime.now().strftime("%Y%m%d")

    # Proveedor (limpiar caracteres especiales)
    proveedor = re.sub(r"[^\w\s]", "", datos.get("proveedor", "Proveedor"))
    proveedor = proveedor.strip().replace(" ", "_")[:30]

    # Descripción en dos palabras (primera línea de concepto)
    concepto = datos.get("lineas", [{}])[0].get("concepto", "")
    palabras = concepto.split()[:2]
    descripcion = "_".join(palabras) if palabras else "factura"
    descripcion = re.sub(r"[^\w]", "_", descripcion)

    nuevo_nombre = f"{fecha_str}_FG_{proveedor}_{descripcion}{ext}"
    ruta_nueva = FACTURAS_DIR / nuevo_nombre

    # Evitar sobreescritura
    contador = 1
    while ruta_nueva.exists():
        nuevo_nombre = f"{fecha_str}_FG_{proveedor}_{descripcion}_{contador}{ext}"
        ruta_nueva = FACTURAS_DIR / nuevo_nombre
        contador += 1

    ruta_temp.rename(ruta_nueva)
    log.info(f"[RENOMBRADO] {ruta_nueva.name}")
    return ruta_nueva


# ─────────────────────────────────────────────
# PASO 5 - GENERAR CSV PARA HOLDED
# ─────────────────────────────────────────────

COLUMNAS_CSV = [
    "Num factura", "Fecha dd/mm/yyyy", "Fecha de vencimiento dd/mm/yyyy",
    "Fecha deducción dd/mm/yyyy", "Descripción", "Nombre del contacto",
    "NIF", "Dirección", "Población", "Código postal", "Provincia", "País",
    "Concepto", "Descripción del producto", "SKU", "Precio unidad", "Unidades",
    "Descuento %", "IVA %", "Retención %", "Inv. Suj. Pasivo (1/0)", "Operación",
    "Cantidad cobrada", "Fecha de cobro", "Cuenta de pago", "Tags separados por -",
    "Nombre cuenta de gasto", "Num. Cuenta de gasto", "Moneda", "Cambio de moneda",
    "Almacén"
]

# Países de la Unión Europea (códigos ISO 3166-1 alpha-2, sin España)
PAISES_UE = {
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR",
    "DE", "GR", "HU", "IE", "IT", "LV", "LT", "LU", "MT", "NL",
    "PL", "PT", "RO", "SK", "SI", "SE"
}

CUENTA_PAGO = "57200002"
NIF_ISBEROAL = "B72598022"  # NIF propio - nunca debe aparecer como proveedor


def detectar_operacion(datos):
    """
    Detecta el tipo de operación fiscal según el país del proveedor:
    - 'general': proveedor español
    - 'intra': proveedor de la UE (intracomunitaria)
    - 'import': proveedor de fuera de la UE (importación)
    """
    codigo_pais = datos.get("codigo_pais", "ES").upper().strip()

    if codigo_pais == "ES":
        return "general"
    elif codigo_pais in PAISES_UE:
        return "intra"
    else:
        return "import"


def datos_a_filas_csv(datos):
    """
    Convierte el diccionario de datos de una factura en filas del CSV de Holded.
    Una factura puede tener múltiples filas (una por línea de concepto).
    """
    filas = []
    lineas = datos.get("lineas", [{"concepto": "Servicio", "precio_unidad": 0, "unidades": 1, "descuento": 0, "iva": 21, "retencion": 0}])

    operacion = detectar_operacion(datos)
    inv_sujeto_pasivo = 1 if datos.get("inversion_sujeto_pasivo", False) else 0

    for i, linea in enumerate(lineas):
        fila = {
            "Num factura":                          datos.get("num_factura", ""),
            "Fecha dd/mm/yyyy":                     datos.get("fecha", ""),
            "Fecha de vencimiento dd/mm/yyyy":      datos.get("fecha_vencimiento", ""),
            "Fecha deducción dd/mm/yyyy":           datos.get("fecha", ""),
            "Descripción":                          f"Factura {datos.get('num_factura', '')}",
            "Nombre del contacto":                  datos.get("proveedor", ""),
            "NIF":                                  datos.get("nif", ""),
            "Dirección":                            datos.get("direccion", ""),
            "Población":                            datos.get("poblacion", ""),
            "Código postal":                        datos.get("codigo_postal", ""),
            "Provincia":                            datos.get("provincia", ""),
            "País":                                 datos.get("pais", "España"),
            "Concepto":                             linea.get("concepto", ""),
            "Descripción del producto":             linea.get("concepto", ""),
            "SKU":                                  "",
            "Precio unidad":                        linea.get("precio_unidad", 0),
            "Unidades":                             linea.get("unidades", 1),
            "Descuento %":                          linea.get("descuento", 0),
            "IVA %":                                linea.get("iva", 21),
            "Retención %":                          linea.get("retencion", 0),
            "Inv. Suj. Pasivo (1/0)":               inv_sujeto_pasivo,
            "Operación":                            operacion,
            # Cantidad cobrada solo en la primera línea para evitar duplicar el total
            "Cantidad cobrada":                     datos.get("total_factura", 0) if i == 0 else "",
            "Fecha de cobro":                       "",
            "Cuenta de pago":                       CUENTA_PAGO,
            "Tags separados por -":                 "",
            "Nombre cuenta de gasto":               "",
            "Num. Cuenta de gasto":                 "",
            "Moneda":                               datos.get("moneda", "EUR").lower(),
            "Cambio de moneda":                     1,
            "Almacén":                              ""
        }
        filas.append(fila)

    return filas


def generar_xlsx(lista_datos, nombre_archivo=None):
    """
    Genera el XLSX de importación para Holded con todas las facturas.
    Formato compatible con la plantilla 'Importar Compras.xlsx' de Holded.
    """
    if not nombre_archivo:
        nombre_archivo = f"importar_compras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    ruta_xlsx = XLSX_OUTPUT_DIR / nombre_archivo
    todas_filas = []

    for datos in lista_datos:
        filas = datos_a_filas_csv(datos)
        todas_filas.extend(filas)

    df = pd.DataFrame(todas_filas, columns=COLUMNAS_CSV)

    with pd.ExcelWriter(ruta_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Compras")
        ws = writer.sheets["Compras"]

        # Formato cabecera
        header_font = Font(bold=True, size=10, name="Arial")
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    log.info(f"[XLSX] Generado: {ruta_xlsx} ({len(todas_filas)} filas, {len(lista_datos)} facturas)")
    return ruta_xlsx


def generar_csv(lista_datos, nombre_archivo=None):
    """
    Genera el CSV de importación para Holded (respaldo/alternativa).
    """
    if not nombre_archivo:
        nombre_archivo = f"importar_compras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    ruta_csv = CSV_OUTPUT_DIR / nombre_archivo
    todas_filas = []

    for datos in lista_datos:
        filas = datos_a_filas_csv(datos)
        todas_filas.extend(filas)

    df = pd.DataFrame(todas_filas, columns=COLUMNAS_CSV)
    df.to_csv(ruta_csv, index=False, encoding="utf-8-sig")

    log.info(f"[CSV] Generado: {ruta_csv} ({len(todas_filas)} filas, {len(lista_datos)} facturas)")
    return ruta_csv


# ─────────────────────────────────────────────
# PASO 6 - IMPORTAR EN HOLDED VIA API
# ─────────────────────────────────────────────

HOLDED_BASE_URL = "https://api.holded.com/api/invoicing/v1"

# Cache de contactos de Holded (NIF -> {id, name})
_contactos_cache = None


def cargar_contactos_holded():
    """
    Carga todos los contactos de Holded y los indexa por NIF (campo 'code').
    La API de Holded NO soporta busqueda por NIF, hay que descargarlos todos.
    """
    global _contactos_cache
    if _contactos_cache is not None:
        return _contactos_cache

    headers = {"key": HOLDED_API_KEY}
    page = 1
    todos = []
    while True:
        resp = requests.get(f"{HOLDED_BASE_URL}/contacts", headers=headers, params={"page": page})
        batch = resp.json()
        if not batch:
            break
        todos.extend(batch)
        if len(batch) < 500:
            break
        page += 1

    _contactos_cache = {}
    for c in todos:
        code = (c.get("code") or "").strip()
        if code:
            _contactos_cache[code] = {"id": c["id"], "name": c.get("name", "")}

    log.info(f"[CONTACTOS] Cargados {len(todos)} contactos, {len(_contactos_cache)} con NIF")
    return _contactos_cache


def buscar_contacto_por_nif(nif):
    """Busca un contacto en Holded por NIF usando el campo 'code'."""
    index = cargar_contactos_holded()
    nif_limpio = nif.replace("-", "").replace(" ", "").strip().upper()
    return index.get(nif_limpio)


def crear_contacto_holded(datos):
    """
    Busca el contacto (proveedor) en Holded por NIF. Si no existe, lo crea.
    """
    headers = {
        "key": HOLDED_API_KEY,
        "Content-Type": "application/json"
    }

    nif = datos.get("nif", "")
    existente = buscar_contacto_por_nif(nif)
    if existente:
        log.info(f"[CONTACTO] Ya existe: {existente['name']} (NIF: {nif})")
        return existente["id"]

    # Crear nuevo contacto con NIF en campo 'code'
    nuevo_contacto = {
        "name":       datos.get("proveedor", ""),
        "code":       nif,
        "vatnumber":  nif,
        "address":    datos.get("direccion", ""),
        "city":       datos.get("poblacion", ""),
        "postalCode": str(datos.get("codigo_postal", "")),
        "province":   datos.get("provincia", ""),
        "country":    "ES",
        "type":       "supplier"
    }

    resp = requests.post(
        f"{HOLDED_BASE_URL}/contacts",
        headers=headers,
        json=nuevo_contacto
    )

    if resp.status_code in [200, 201]:
        contacto_id = resp.json().get("id")
        log.info(f"[CONTACTO] Creado: {datos.get('proveedor')} (NIF: {nif}, ID: {contacto_id})")
        # Actualizar cache
        global _contactos_cache
        if _contactos_cache is not None:
            _contactos_cache[nif] = {"id": contacto_id, "name": datos.get("proveedor", "")}
        return contacto_id
    else:
        log.error(f"[ERROR] Creando contacto: {resp.text}")
        return None


def crear_factura_holded(datos, contacto_id):
    """
    Crea la factura de compra en Holded.
    """
    headers = {
        "key": HOLDED_API_KEY,
        "Content-Type": "application/json"
    }

    # Convertir fecha a timestamp Unix
    def fecha_a_timestamp(fecha_str):
        try:
            dt = datetime.strptime(fecha_str, "%d/%m/%Y")
            return int(dt.timestamp())
        except Exception:
            return int(datetime.now().timestamp())

    # Preparar lineas de la factura
    # NOTA: Holded usa 'subtotal' (no 'price') en items, y 'invoiceNum' (no 'docNumber')
    items = []
    for linea in datos.get("lineas", []):
        item = {
            "name":      linea.get("concepto", ""),
            "units":     linea.get("unidades", 1),
            "subtotal":  linea.get("precio_unidad", 0),
            "discount":  linea.get("descuento", 0),
            "tax":       linea.get("iva", 21),
        }
        retencion = linea.get("retencion", 0)
        if retencion:
            item["retention"] = retencion
        items.append(item)

    payload = {
        "contactId":   contacto_id,
        "invoiceNum":  datos.get("num_factura", ""),
        "date":        fecha_a_timestamp(datos.get("fecha", "")),
        "dueDate":     fecha_a_timestamp(datos.get("fecha_vencimiento", "")),
        "notes":       f"Importado automaticamente - {datos.get('num_factura')}",
        "items":       items,
        "currency":    datos.get("moneda", "EUR").lower(),
    }

    if datos.get("inversion_sujeto_pasivo", False):
        payload["reverseCharge"] = True

    resp = requests.post(
        f"{HOLDED_BASE_URL}/documents/purchase",
        headers=headers,
        json=payload
    )

    if resp.status_code in [200, 201]:
        factura_id = resp.json().get("id")
        log.info(f"[HOLDED] Factura creada: {datos.get('num_factura')} (ID: {factura_id})")
        return factura_id
    else:
        log.error(f"[ERROR] Creando factura en Holded: {resp.status_code} - {resp.text}")
        return None


def importar_facturas_holded(lista_datos):
    """
    Importa todas las facturas en Holded una a una.
    """
    resultados = []
    for datos in lista_datos:
        try:
            # Crear o recuperar contacto
            contacto_id = crear_contacto_holded(datos)
            if not contacto_id:
                log.warning(f"[SKIP] Factura {datos.get('num_factura')} - no se pudo crear contacto")
                continue

            # Crear factura
            factura_id = crear_factura_holded(datos, contacto_id)
            resultados.append({
                "num_factura":  datos.get("num_factura"),
                "proveedor":    datos.get("proveedor"),
                "total":        datos.get("total_factura"),
                "holded_id":    factura_id,
                "estado":       "OK" if factura_id else "ERROR"
            })

            time.sleep(0.5)  # Respetar rate limits de la API

        except Exception as e:
            log.error(f"[ERROR] Procesando factura {datos.get('num_factura', '?')}: {e}")
            resultados.append({
                "num_factura": datos.get("num_factura"),
                "estado":      f"ERROR: {e}"
            })

    return resultados


# ─────────────────────────────────────────────
# CONTROL DE DUPLICADOS
# ─────────────────────────────────────────────

def cargar_procesados():
    """Carga el registro de message_ids ya procesados."""
    if PROCESADOS_PATH.exists():
        with open(PROCESADOS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"mensajes": {}}


def guardar_procesados(registro):
    """Guarda el registro de message_ids procesados."""
    with open(PROCESADOS_PATH, "w", encoding="utf-8") as f:
        json.dump(registro, f, indent=2, ensure_ascii=False)


def marcar_procesado(registro, msg_id, facturas_ids):
    """Marca un mensaje como procesado con las facturas que contenía."""
    registro["mensajes"][msg_id] = {
        "fecha_proceso": datetime.now().isoformat(),
        "facturas": facturas_ids
    }


def ya_procesado(registro, msg_id):
    """Comprueba si un mensaje ya fue procesado."""
    return msg_id in registro.get("mensajes", {})


# ─────────────────────────────────────────────
# ORQUESTADOR PRINCIPAL
# ─────────────────────────────────────────────

def ejecutar_agente(dias_atras=7, solo_xlsx=False, importar_holded=True, forzar=False):
    """
    Ejecuta el agente completo de registro de facturas.

    Args:
        dias_atras:       Dias hacia atras para buscar correos (por defecto 7)
        solo_xlsx:        Si True, solo genera el XLSX sin importar en Holded
        importar_holded:  Si True, importa directamente en Holded via API
        forzar:           Si True, reprocesa mensajes ya procesados (ignora duplicados)
    """
    log.info("=" * 60)
    log.info("INICIANDO AGENTE DE FACTURAS DE COMPRA - ISBEROAL")
    log.info("=" * 60)

    # SHADOW MODE: si la variable de entorno está activa, forzar modo seguro.
    # Útil para validación en paralelo al agente local en producción.
    if os.getenv("ISBEROAL_SHADOW_MODE", "").lower() == "true":
        log.warning("[SHADOW MODE] Activo. No se importará a Holded. Solo se generará XLSX.")
        solo_xlsx = True
        importar_holded = False

    # Cargar registro de duplicados
    registro = cargar_procesados()

    # PASO 1 - Conectar Gmail
    service = conectar_gmail()

    # Obtener correos
    mensajes = obtener_correos_con_facturas(service, dias_atras=dias_atras)
    if not mensajes:
        log.info("No hay correos nuevos con facturas. Finalizando.")
        return

    # Filtrar duplicados
    if not forzar:
        mensajes_nuevos = [m for m in mensajes if not ya_procesado(registro, m["id"])]
        saltados = len(mensajes) - len(mensajes_nuevos)
        if saltados > 0:
            log.info(f"[DUPLICADOS] {saltados} mensajes ya procesados (saltados)")
        mensajes = mensajes_nuevos

    if not mensajes:
        log.info("Todos los correos ya fueron procesados. Finalizando.")
        return

    log.info(f"[NUEVOS] {len(mensajes)} mensajes por procesar")

    lista_datos = []
    archivos_procesados = []
    nums_factura_vistos = set()  # Evitar duplicados de num_factura en el mismo lote

    for msg in mensajes:
        msg_id = msg["id"]
        log.info(f"\nProcesando mensaje: {msg_id}")
        facturas_en_mensaje = []

        try:
            # PASO 2 - Descargar adjuntos
            adjuntos = descargar_adjuntos(service, msg_id)

            for adjunto in adjuntos:
                ruta_temp = adjunto["ruta"]

                try:
                    # PASO 3 - Extraer datos con OCR + IA
                    datos = extraer_datos_factura(ruta_temp)

                    # VALIDACION: comprobar que es factura valida
                    valido, motivo = validar_datos_factura(datos, ruta_temp)
                    if not valido:
                        log.warning(f"[DESCARTADO] {ruta_temp.name}: {motivo}")
                        ruta_temp.rename(FACTURAS_DIR / f"DESCARTADO_{ruta_temp.name}")
                        continue

                    # Evitar duplicados de num_factura en el mismo lote
                    num_fac = datos.get("num_factura", "")
                    if num_fac in nums_factura_vistos:
                        log.warning(f"[DUPLICADO] Factura {num_fac} ya procesada en este lote, saltando")
                        ruta_temp.rename(FACTURAS_DIR / f"DUP_{ruta_temp.name}")
                        continue
                    nums_factura_vistos.add(num_fac)

                    # PASO 2 (continuacion) - Renombrar archivo
                    ruta_final = renombrar_factura(ruta_temp, datos)
                    archivos_procesados.append(ruta_final)

                    lista_datos.append(datos)
                    facturas_en_mensaje.append(datos.get("num_factura", "?"))

                except json.JSONDecodeError as e:
                    log.error(f"Error parseando respuesta de Claude para {ruta_temp.name}: {e}")
                    ruta_temp.rename(FACTURAS_DIR / f"ERROR_{ruta_temp.name}")

                except Exception as e:
                    log.error(f"Error procesando adjunto {ruta_temp.name}: {e}")

            # Marcar mensaje como procesado (aunque haya errores parciales)
            marcar_procesado(registro, msg_id, facturas_en_mensaje)
            guardar_procesados(registro)

        except Exception as e:
            log.error(f"Error descargando adjuntos del mensaje {msg_id}: {e}")

    if not lista_datos:
        log.warning("No se pudieron extraer datos de ninguna factura.")
        return

    # PASO 5 - Generar XLSX (formato principal) y CSV (respaldo)
    ruta_xlsx = generar_xlsx(lista_datos)
    ruta_csv = generar_csv(lista_datos)
    log.info(f"\n[XLSX] Listo para importar: {ruta_xlsx}")

    # PASO 6 - Importar en Holded
    if importar_holded and not solo_xlsx:
        log.info("\nImportando facturas en Holded...")
        resultados = importar_facturas_holded(lista_datos)

        log.info("\n" + "=" * 60)
        log.info("RESUMEN FINAL")
        log.info("=" * 60)
        ok    = sum(1 for r in resultados if r.get("estado") == "OK")
        error = sum(1 for r in resultados if r.get("estado") != "OK")
        log.info(f"Facturas importadas: {ok}")
        log.info(f"Errores: {error}")
        for r in resultados:
            estado = "OK" if r.get("estado") == "OK" else "ERROR"
            log.info(f"  [{estado}] {r.get('num_factura')} | {r.get('proveedor')} | {r.get('total')} EUR")
    else:
        log.info("\nModo solo XLSX. Importacion manual desde Holded con el archivo generado.")

    log.info("\nAgente finalizado.")
    return lista_datos, ruta_xlsx


# ─────────────────────────────────────────────
# PUNTO DE ENTRADA
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Agente de Facturas de Compra - ISBEROAL")
    parser.add_argument("--dias",       type=int,  default=7,  help="Dias hacia atras para buscar correos")
    parser.add_argument("--solo-xlsx",  action="store_true",   help="Generar solo el XLSX sin importar en Holded")
    parser.add_argument("--no-holded",  action="store_true",   help="No importar en Holded")
    parser.add_argument("--forzar",     action="store_true",   help="Reprocesar mensajes ya procesados (ignorar duplicados)")
    args = parser.parse_args()

    ejecutar_agente(
        dias_atras=args.dias,
        solo_xlsx=args.solo_xlsx,
        importar_holded=not args.no_holded,
        forzar=args.forzar
    )
