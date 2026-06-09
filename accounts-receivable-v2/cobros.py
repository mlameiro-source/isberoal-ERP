"""
Agente de cobros (accounts receivable) - ISBEROAL.

Lee las facturas de venta de Holded (emitidas + borradores) mediante el modulo
holded_ventas y genera un XLSX a nivel de factura con la situacion de cobro.
Read-only contra Holded: solo lee, nunca escribe.

Uso (en local, PowerShell):
    pip install -r requirements.txt
    python cobros.py
"""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from holded_ventas import leer_facturas_venta

# Ruta de salida. Por defecto el directorio actual. Para dejarlo en la unidad
# compartida de Google, define COBROS_OUTPUT en el .env con la ruta G:\...
SALIDA = os.environ.get("COBROS_OUTPUT") or "cobros_isberoal.xlsx"

ENCABEZADOS = [
    "ID Holded", "Estado", "Nº factura", "Cliente",
    "Fecha emisión", "Fecha vencimiento", "Descripción",
    "Base", "IVA", "Total", "Cobrado", "Pendiente",
    "Situación", "Moneda", "Tags",
]

_AZUL = "1F3864"
_COLOR_SIT = {
    "Cobrada": "C6EFCE",
    "Parcial": "FFEB9C",
    "Vencida": "FFC7CE",
    "Pendiente": "FFF2CC",
    "Prevista": "E7E6E6",
}


def situacion(f, hoy):
    if f["es_borrador"]:
        return "Prevista"
    if f["pendiente"] <= 0.005:
        return "Cobrada"
    if f["fecha_vencimiento"] and f["fecha_vencimiento"] < hoy:
        return "Vencida"
    if f["cobrado"] > 0.005:
        return "Parcial"
    return "Pendiente"


def _formato_cobros(ws, n_filas):
    cab_fill = PatternFill("solid", fgColor=_AZUL)
    cab_font = Font(name="Arial", bold=True, color="FFFFFF")
    for col in range(1, len(ENCABEZADOS) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = cab_fill
        c.font = cab_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    cuerpo = Font(name="Arial")
    for row in range(2, n_filas + 2):
        for col in range(1, len(ENCABEZADOS) + 1):
            ws.cell(row=row, column=col).font = cuerpo
        for col in (5, 6):  # fechas
            ws.cell(row=row, column=col).number_format = "DD/MM/YYYY"
        for col in (8, 9, 10, 11, 12):  # importes
            ws.cell(row=row, column=col).number_format = "#,##0.00"
        sit = ws.cell(row=row, column=13)
        color = _COLOR_SIT.get(sit.value)
        if color:
            sit.fill = PatternFill("solid", fgColor=color)

    anchos = [22, 10, 12, 26, 14, 16, 30, 12, 12, 13, 12, 13, 12, 8, 16]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    ws.freeze_panes = "A2"
    if n_filas > 0:
        ws.auto_filter.ref = f"A1:O{n_filas + 1}"


def _hoja_resumen(wb, n_filas):
    ws = wb.create_sheet("Resumen", 0)
    titulo = Font(name="Arial", bold=True, size=12)
    etiqueta = Font(name="Arial")
    valor = Font(name="Arial", bold=True)

    ws["A1"] = "Resumen de cobros - ISBEROAL"
    ws["A1"].font = titulo
    ws["A2"] = "Actualizado:"
    ws["A2"].font = etiqueta
    ws["B2"] = date.today()
    ws["B2"].number_format = "DD/MM/YYYY"
    ws["B2"].font = valor

    filas = [
        ("Facturas emitidas (nº)", '=COUNTIF(Cobros!B:B,"Emitida")', "0"),
        ("Borradores / previstos (nº)", '=COUNTIF(Cobros!B:B,"Borrador")', "0"),
        ("Total emitido", '=SUMIF(Cobros!B:B,"Emitida",Cobros!J:J)', "#,##0.00"),
        ("Cobrado", '=SUMIF(Cobros!B:B,"Emitida",Cobros!K:K)', "#,##0.00"),
        ("Pendiente de cobro (emitidas)", '=SUMIF(Cobros!B:B,"Emitida",Cobros!L:L)', "#,##0.00"),
        ("De ello, VENCIDO", '=SUMIF(Cobros!M:M,"Vencida",Cobros!L:L)', "#,##0.00"),
        ("Previsto en borradores (total)", '=SUMIF(Cobros!B:B,"Borrador",Cobros!J:J)', "#,##0.00"),
    ]
    r = 4
    for et, fx, fmt in filas:
        ws.cell(row=r, column=1, value=et).font = etiqueta
        cel = ws.cell(row=r, column=2, value=fx)
        cel.font = valor
        cel.number_format = fmt
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16


def construir(facturas, ruta):
    hoy = date.today()
    facturas = sorted(
        facturas,
        key=lambda x: (x["fecha_emision"] or date.max, x["num_factura"] or ""),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cobros"
    ws.append(ENCABEZADOS)

    for f in facturas:
        ws.append([
            f["id"],
            "Borrador" if f["es_borrador"] else "Emitida",
            f["num_factura"] or "—",
            f["cliente"],
            f["fecha_emision"],
            f["fecha_vencimiento"],
            f["descripcion"],
            f["base"], f["iva"], f["total"], f["cobrado"], f["pendiente"],
            situacion(f, hoy),
            f["moneda"],
            " - ".join(f["tags"]) if f["tags"] else "",
        ])

    _formato_cobros(ws, len(facturas))
    _hoja_resumen(wb, len(facturas))
    wb.save(ruta)
    return ruta


if __name__ == "__main__":
    facturas = leer_facturas_venta()
    ruta = construir(facturas, SALIDA)
    emitidas = sum(1 for f in facturas if not f["es_borrador"])
    borradores = len(facturas) - emitidas
    print(f"Generado: {os.path.abspath(ruta)}")
    print(f"  {emitidas} emitidas, {borradores} borradores")