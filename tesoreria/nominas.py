"""
Modulo de nominas para la proyeccion de tesoreria - ISBEROAL.

Lee nominas_datos.xlsx y devuelve el calendario de salidas de caja por
personal. El archivo de datos tiene UNA FILA POR EMPLEADO Y MES (el modulo
agrega por mes el solo); en fase 2 lo rellenara el OCR de los PDF de la
gestoria, que produce exactamente ese formato (un registro por nomina).

Salidas de caja generadas:
  - Neto del mes M        -> ultimo dia habil de M menos 2 dias habiles
  - TGSS devengada en M   -> ultimo dia habil de M+1
  - IRPF retenido         -> acumulado por trimestre, dia 20 del mes siguiente
                             al cierre (modelo 111); si cae en finde, siguiente habil

No toca Holded ni ningun servicio externo. Solo lee el XLSX local.

Hoja "Meses": mes | bruto_total | ss_empresa | ss_trabajador | irpf | tipo | nota
  - mes: fecha o texto YYYY-MM (si es fecha, el dia se ignora)
  - tipo: real / prevision (si un mes mezcla ambos, el mes cuenta como prevision)
  - nota: libre (tipicamente el nombre del empleado)
Hoja "Ajustes": desde_mes | bruto | ss_empresa | ss_trabajador | irpf | nota
  - deltas que se suman a los meses ESTIMADOS (sin fila propia) desde desde_mes

Uso:
    python nominas.py --plantilla         crea nominas_datos.xlsx vacio
    python nominas.py                     salidas de caja, 6 meses vista
    python nominas.py --hasta 2026-12     salidas de caja hasta ese mes
"""

import argparse
import calendar
import os
from datetime import date, timedelta

from openpyxl import Workbook, load_workbook

ARCHIVO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "nominas_datos.xlsx")

# Regla de pago de nomina: ultimo dia habil del mes menos N dias habiles
DIAS_HABILES_ANTES_NOMINA = 2

COLS_MESES = ["mes", "bruto_total", "ss_empresa", "ss_trabajador", "irpf", "tipo", "nota"]
COLS_AJUSTES = ["desde_mes", "bruto", "ss_empresa", "ss_trabajador", "irpf", "nota"]


# ---------- calendario ----------

def _es_habil(d):
    return d.weekday() < 5  # lunes-viernes; festivos no contemplados


def _ultimo_habil_mes(anio, mes):
    d = date(anio, mes, calendar.monthrange(anio, mes)[1])
    while not _es_habil(d):
        d -= timedelta(days=1)
    return d


def _restar_habiles(d, n):
    while n > 0:
        d -= timedelta(days=1)
        if _es_habil(d):
            n -= 1
    return d


def _siguiente_habil(d):
    while not _es_habil(d):
        d += timedelta(days=1)
    return d


def _mes_siguiente(anio, mes):
    return (anio + 1, 1) if mes == 12 else (anio, mes + 1)


def fecha_pago_nomina(anio, mes):
    return _restar_habiles(_ultimo_habil_mes(anio, mes), DIAS_HABILES_ANTES_NOMINA)


def fecha_pago_tgss(anio, mes):
    """La SS devengada en (anio, mes) se carga a fin del mes siguiente."""
    a2, m2 = _mes_siguiente(anio, mes)
    return _ultimo_habil_mes(a2, m2)


def fecha_pago_irpf(anio, trimestre):
    """Modelo 111: dia 20 del mes siguiente al cierre del trimestre."""
    mes_pago = {1: 4, 2: 7, 3: 10, 4: 1}[trimestre]
    anio_pago = anio + 1 if trimestre == 4 else anio
    return _siguiente_habil(date(anio_pago, mes_pago, 20))


# ---------- lectura del archivo de datos ----------

def _parse_mes(v, fila_n, hoja):
    """Acepta '2026-06' (texto) o una celda de fecha. Devuelve (anio, mes)."""
    if hasattr(v, "year"):
        return (v.year, v.month)
    partes = str(v).strip().split("-")
    if len(partes) < 2 or not partes[0].isdigit() or not partes[1].isdigit():
        raise ValueError(
            f"Hoja {hoja}, fila {fila_n}: mes invalido {v!r}. "
            "Formato esperado: YYYY-MM o una fecha."
        )
    return (int(partes[0]), int(partes[1]))


def _num_estricto(v, fila_n, columna):
    """Importe obligatorio y numerico. Aborta con error claro si no lo es,
    para que una celda vacia o con texto nunca se convierta en 0 en silencio."""
    if not isinstance(v, (int, float)):
        raise ValueError(
            f"Hoja Meses, fila {fila_n}, columna {columna}: valor no numerico "
            f"({v!r}). Revisa el archivo: importes vacios o con texto no se "
            "aceptan para no falsear la proyeccion."
        )
    return round(float(v), 2)


def leer_datos(ruta=ARCHIVO):
    if not os.path.exists(ruta):
        raise FileNotFoundError(
            f"No existe {ruta}. Crealo con: python nominas.py --plantilla"
        )
    wb = load_workbook(ruta, data_only=True)
    if "Meses" not in wb.sheetnames:
        raise ValueError(
            f"El archivo no tiene hoja 'Meses' (hojas: {wb.sheetnames}). "
            "Renombra la hoja de datos a 'Meses'."
        )

    # Agregacion: una fila por empleado y mes -> totales por mes
    meses = {}
    fila_n = 1
    for fila in wb["Meses"].iter_rows(min_row=2, values_only=True):
        fila_n += 1
        if not fila or fila[0] is None:
            continue
        clave = _parse_mes(fila[0], fila_n, "Meses")
        tipo = (str(fila[5]).strip().lower() if len(fila) > 5 and fila[5] else "real")
        if tipo not in ("real", "prevision"):
            raise ValueError(
                f"Hoja Meses, fila {fila_n}: tipo invalido {fila[5]!r}. "
                "Valores aceptados: real, prevision."
            )
        acc = meses.setdefault(clave, {
            "bruto": 0.0, "ss_empresa": 0.0, "ss_trabajador": 0.0,
            "irpf": 0.0, "tipos": set(), "filas": 0,
        })
        acc["bruto"] += _num_estricto(fila[1], fila_n, "bruto_total")
        acc["ss_empresa"] += _num_estricto(fila[2], fila_n, "ss_empresa")
        acc["ss_trabajador"] += _num_estricto(fila[3], fila_n, "ss_trabajador")
        acc["irpf"] += _num_estricto(fila[4], fila_n, "irpf")
        acc["tipos"].add(tipo)
        acc["filas"] += 1

    for clave, d in meses.items():
        for k in ("bruto", "ss_empresa", "ss_trabajador", "irpf"):
            d[k] = round(d[k], 2)
        d["tipo"] = "real" if d.pop("tipos") == {"real"} else "prevision"

    ajustes = []
    if "Ajustes" in wb.sheetnames:
        fila_n = 1
        for fila in wb["Ajustes"].iter_rows(min_row=2, values_only=True):
            fila_n += 1
            if not fila or fila[0] is None:
                continue
            ajustes.append({
                "desde": _parse_mes(fila[0], fila_n, "Ajustes"),
                "bruto": round(float(fila[1] or 0), 2),
                "ss_empresa": round(float(fila[2] or 0), 2),
                "ss_trabajador": round(float(fila[3] or 0), 2),
                "irpf": round(float(fila[4] or 0), 2),
            })

    if not meses:
        raise ValueError("La hoja Meses esta vacia. Rellena al menos un mes real.")
    return meses, ajustes


def _datos_mes(clave, meses, ajustes):
    """Datos del mes pedido. Si no hay filas de ese mes, estima: ultimo mes
    conocido anterior + ajustes con desde_mes <= mes pedido y > ese conocido."""
    if clave in meses:
        d = meses[clave]
        return {
            "bruto": d["bruto"], "ss_empresa": d["ss_empresa"],
            "ss_trabajador": d["ss_trabajador"], "irpf": d["irpf"],
            "origen": d["tipo"],
        }

    anteriores = [k for k in meses if k <= clave]
    if not anteriores:
        return None  # mes anterior al primer dato: no hay base para estimar
    base_clave = max(anteriores)
    base = meses[base_clave]
    d = {
        "bruto": base["bruto"],
        "ss_empresa": base["ss_empresa"],
        "ss_trabajador": base["ss_trabajador"],
        "irpf": base["irpf"],
        "origen": "estimado",
    }
    for a in ajustes:
        if base_clave < a["desde"] <= clave:
            d["bruto"] += a["bruto"]
            d["ss_empresa"] += a["ss_empresa"]
            d["ss_trabajador"] += a["ss_trabajador"]
            d["irpf"] += a["irpf"]
    for k in ("bruto", "ss_empresa", "ss_trabajador", "irpf"):
        d[k] = round(d[k], 2)
    return d


# ---------- API principal (la consumira proyeccion.py) ----------

def salidas_caja(hasta, desde=None, ruta=ARCHIVO):
    """
    Lista de salidas de caja por personal entre 'desde' (defecto: hoy) y el
    final del mes 'hasta' (tupla (anio, mes)). Cada salida es un dict:
    {fecha, concepto, importe, mes_devengo, origen}
    """
    meses, ajustes = leer_datos(ruta)
    desde = desde or date.today()
    primero = min(meses)

    salidas = []
    irpf_trimestres = {}

    a, m = primero
    while (a, m) <= hasta:
        d = _datos_mes((a, m), meses, ajustes)
        if d:
            neto = round(d["bruto"] - d["ss_trabajador"] - d["irpf"], 2)
            tgss = round(d["ss_empresa"] + d["ss_trabajador"], 2)
            salidas.append({
                "fecha": fecha_pago_nomina(a, m), "concepto": "Nomina neta",
                "importe": neto, "mes_devengo": f"{a}-{m:02d}", "origen": d["origen"],
            })
            salidas.append({
                "fecha": fecha_pago_tgss(a, m), "concepto": "Seguros Sociales",
                "importe": tgss, "mes_devengo": f"{a}-{m:02d}", "origen": d["origen"],
            })
            t = (a, (m - 1) // 3 + 1)
            acc = irpf_trimestres.setdefault(t, {"importe": 0.0, "origenes": set()})
            acc["importe"] += d["irpf"]
            acc["origenes"].add(d["origen"])
        a, m = _mes_siguiente(a, m)

    for (anio_t, tri), acc in irpf_trimestres.items():
        origen = "real" if acc["origenes"] == {"real"} else (
            "prevision" if "estimado" not in acc["origenes"] else "estimado")
        salidas.append({
            "fecha": fecha_pago_irpf(anio_t, tri), "concepto": "IRPF (mod. 111)",
            "importe": round(acc["importe"], 2),
            "mes_devengo": f"{anio_t}-T{tri}", "origen": origen,
        })

    fin_mes = date(hasta[0], hasta[1], calendar.monthrange(hasta[0], hasta[1])[1])
    salidas = [s for s in salidas if desde <= s["fecha"] <= fin_mes]
    salidas.sort(key=lambda s: s["fecha"])
    return salidas


# ---------- utilidades de linea de comandos ----------

def crear_plantilla(ruta=ARCHIVO):
    if os.path.exists(ruta):
        raise FileExistsError(f"Ya existe {ruta}. No lo sobrescribo.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Meses"
    ws.append(COLS_MESES)
    ws2 = wb.create_sheet("Ajustes")
    ws2.append(COLS_AJUSTES)
    wb.save(ruta)
    return ruta


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--plantilla", action="store_true", help="crear nominas_datos.xlsx vacio")
    p.add_argument("--hasta", help="mes final YYYY-MM (defecto: 6 meses vista)")
    args = p.parse_args()

    if args.plantilla:
        print(f"Plantilla creada: {crear_plantilla()}")
        print("Rellena la hoja Meses (una fila por empleado y mes) y vuelve a ejecutar.")
        raise SystemExit(0)

    if args.hasta:
        hasta = tuple(int(x) for x in args.hasta.split("-")[:2])
    else:
        hoy = date.today()
        a, m = hoy.year, hoy.month
        for _ in range(6):
            a, m = _mes_siguiente(a, m)
        hasta = (a, m)

    lista = salidas_caja(hasta)
    if not lista:
        print("Sin salidas de caja en el rango. Revisa los datos.")
        raise SystemExit(0)

    print(f"{'Fecha':<12} {'Concepto':<18} {'Importe':>14} {'Devengo':<9} Origen")
    for s in lista:
        print(f"{s['fecha'].strftime('%d/%m/%Y'):<12} {s['concepto']:<18} "
              f"{s['importe']:>14,.2f} {s['mes_devengo']:<9} {s['origen']}")
    print(f"\nTotal salidas en el rango: {sum(s['importe'] for s in lista):,.2f}")